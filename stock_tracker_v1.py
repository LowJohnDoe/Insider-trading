"""
Stock Tracker v1

Goal:
- Pull recent 13F-HR filings for one or more managers
- Find the filing detail page
- Locate the information table XML when available
- Parse holdings rows (issuer, class, cusip, value, shares, discretion)

Notes:
- This is a starter version. It does NOT yet do:
    * CUSIP -> ticker mapping
    * price filtering for "cheap" stocks
    * Google Sheets output
    * Twilio SMS alerts
- It is written to be easy to debug step by step.
- Before sharing publicly, replace placeholder contact info with a real project email you control.
"""

from __future__ import annotations

import time
import logging
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import List, Optional, Dict, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# =========================
# Configuration
# =========================

USER_AGENT = "StockTrackerTool/0.1 (educational project; contact: StToolv1@aol.com)"
REQUEST_DELAY_SECONDS = 0.5
BASE_SEC = "https://www.sec.gov"

# Replace these later with the managers you want to track.
# These are example CIKs only.
MANAGERS = [
    {"name": "Berkshire Hathaway", "cik": "0001067983"},
    {"name": "Scion Asset Management", "cik": "0001649339"},
]

HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Encoding": "gzip, deflate",
    "Host": "www.sec.gov",
}


# =========================
# Data models
# =========================

@dataclass
class Holding:
    manager_name: str
    manager_cik: str
    filing_date: str
    report_period: str
    issuer: str
    title_of_class: str
    cusip: str
    value: Optional[float]
    shares_or_prn_amount: Optional[float]
    shares_or_prn_type: str
    investment_discretion: str
    filing_detail_url: str
    info_table_url: str


# New AggregatedHolding dataclass
@dataclass
class AggregatedHolding:
    manager_name: str
    manager_cik: str
    report_period: str
    filing_date: str
    issuer: str
    title_of_class: str
    cusip: str
    total_value: float
    total_shares: float
    implied_price: Optional[float]
    row_count: int


# =========================
# Helpers
# =========================

def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    return session


def safe_get(session: requests.Session, url: str, timeout: int = 30) -> requests.Response:
    """GET with a polite delay and basic error handling."""
    time.sleep(REQUEST_DELAY_SECONDS)
    response = session.get(url, timeout=timeout)
    response.raise_for_status()
    return response


def text_or_empty(node: Optional[ET.Element]) -> str:
    return node.text.strip() if node is not None and node.text else ""


def parse_float(value: str) -> Optional[float]:
    value = value.replace(",", "").strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


# Helper to compute implied price
def compute_implied_price(value: Optional[float], shares: Optional[float]) -> Optional[float]:
    if value is None or shares is None or shares <= 0:
        return None
    return value / shares


# =========================
# SEC lookup functions
# =========================

def get_company_filings_url(cik: str, filing_type: str = "13F-HR", count: int = 100) -> str:
    return (
        f"{BASE_SEC}/cgi-bin/browse-edgar?action=getcompany"
        f"&CIK={cik}&type={filing_type}&owner=exclude&count={count}"
    )


def get_recent_filing_detail_urls(session: requests.Session, cik: str, max_filings: int = 5) -> List[str]:
    """
    From the company's filings page, collect recent filing detail page URLs.
    """
    url = get_company_filings_url(cik)
    response = safe_get(session, url)
    soup = BeautifulSoup(response.text, "html.parser")

    detail_urls: List[str] = []

    table = soup.find("table", class_="tableFile2")
    if table is None:
        logging.warning("Could not find filings table for CIK %s", cik)
        return detail_urls

    rows = table.find_all("tr")
    for row in rows[1:]:
        cols = row.find_all("td")
        if len(cols) < 2:
            continue

        filing_type = cols[0].get_text(strip=True)
        if filing_type != "13F-HR":
            continue

        documents_cell = cols[1]
        link = documents_cell.find("a")
        if link and link.get("href"):
            detail_urls.append(urljoin(BASE_SEC, link["href"]))

        if len(detail_urls) >= max_filings:
            break

    return detail_urls


def parse_filing_detail_page(session: requests.Session, detail_url: str) -> dict:
    """
    Open the filing detail page and find:
    - filing date
    - report period
    - the actual 13F information table XML URL
    """
    response = safe_get(session, detail_url)
    soup = BeautifulSoup(response.text, "html.parser")

    filing_date = ""
    report_period = ""
    info_table_url = ""

    info_heads = soup.find_all("div", class_="infoHead")
    info_values = soup.find_all("div", class_="info")
    for head, value in zip(info_heads, info_values):
        head_text = head.get_text(" ", strip=True).lower()
        value_text = value.get_text(" ", strip=True)
        if "filing date" in head_text:
            filing_date = value_text
        elif "period of report" in head_text:
            report_period = value_text

    table = soup.find("table", class_="tableFile", summary="Document Format Files")
    if table:
        candidates = []

        for row in table.find_all("tr")[1:]:
            cols = row.find_all("td")
            if len(cols) < 4:
                continue

            seq = cols[0].get_text(" ", strip=True)
            description = cols[1].get_text(" ", strip=True).lower()
            document_cell = cols[2]
            doc_type = cols[3].get_text(" ", strip=True).lower()

            link = document_cell.find("a")
            if not link or not link.get("href"):
                continue

            href = urljoin(BASE_SEC, link["href"])
            filename = link.get_text(" ", strip=True).lower()

            # SEC sometimes routes XML links through an XSL-rendered folder.
            # Convert that viewer path back to the raw XML file path.
            if "/xslForm13F_" in href:
                parts = href.split("/")
                xsl_index = None
                for i, part in enumerate(parts):
                    if part.startswith("xslForm13F_"):
                        xsl_index = i
                        break

                if xsl_index is not None and parts:
                    raw_filename = parts[-1]
                    parts = parts[:xsl_index] + [raw_filename]
                    href = "/".join(parts)

            score = 0

            if href.endswith(".xml"):
                score += 4
            if filename.endswith(".xml"):
                score += 4
            if doc_type == "xml":
                score += 2
            if "information table" in description:
                score += 6
            if "infotable" in filename:
                score += 6
            if "13f" in filename and href.endswith(".xml"):
                score += 3

            if "/xslForm13F_" in href:
                score -= 10
            if filename.endswith(".html"):
                score -= 4
            if "primary_doc" in filename:
                score -= 6
            if "xslf345x" in filename:
                score -= 6
            if "complete submission text file" in description:
                score -= 8

            candidates.append((score, href, description, filename, doc_type, seq))

        candidates.sort(reverse=True, key=lambda x: x[0])

        if candidates and candidates[0][0] > 0:
            best = candidates[0]
            info_table_url = best[1]
            print(f"\nChosen XML: {info_table_url}")
            print(f"Description: {best[2]}")
            print(f"Filename: {best[3]}")
            print(f"Doc type: {best[4]}\n")

    return {
        "filing_date": filing_date,
        "report_period": report_period,
        "info_table_url": info_table_url,
        "detail_url": detail_url,
    }


# =========================
# XML parsing
# =========================

def parse_13f_info_table_xml(
    xml_text: str,
    manager_name: str,
    manager_cik: str,
    filing_date: str,
    report_period: str,
    detail_url: str,
    info_table_url: str,
) -> List[Holding]:
    """
    Parse 13F information table XML into a list of Holding objects.
    """
    holdings: List[Holding] = []

    # Quick sanity check so we don't try parsing HTML as XML holdings
    preview = xml_text[:1000].lower()
    if "<html" in preview or "<html " in preview or "<!doctype html" in preview:
        raise ValueError(f"Selected file is HTML, not holdings XML: {info_table_url}")
    if "informationtable" not in preview and "infotable" not in preview:
        raise ValueError(f"Selected file does not appear to be a 13F holdings XML: {info_table_url}")

    root = ET.fromstring(xml_text)
    
    # Namespaces vary, so ignore them by checking element suffixes.
    for info_table in root.iter():
        if not info_table.tag.endswith("infoTable"):
            continue

        children = list(info_table)
        lookup = {}
        for child in children:
            tag_name = child.tag.split("}")[-1]
            lookup[tag_name] = child

        ssh_prnamt_node = None
        ssh_prnamt_type_node = None
        shrs_or_prn_amt = lookup.get("shrsOrPrnAmt")
        if shrs_or_prn_amt is not None:
            for sub in list(shrs_or_prn_amt):
                sub_name = sub.tag.split("}")[-1]
                if sub_name == "sshPrnamt":
                    ssh_prnamt_node = sub
                elif sub_name == "sshPrnamtType":
                    ssh_prnamt_type_node = sub

        holdings.append(
            Holding(
                manager_name=manager_name,
                manager_cik=manager_cik,
                filing_date=filing_date,
                report_period=report_period,
                issuer=text_or_empty(lookup.get("nameOfIssuer")),
                title_of_class=text_or_empty(lookup.get("titleOfClass")),
                cusip=text_or_empty(lookup.get("cusip")),
                value=parse_float(text_or_empty(lookup.get("value"))),
                shares_or_prn_amount=parse_float(text_or_empty(ssh_prnamt_node)),
                shares_or_prn_type=text_or_empty(ssh_prnamt_type_node),
                investment_discretion=text_or_empty(lookup.get("investmentDiscretion")),
                filing_detail_url=detail_url,
                info_table_url=info_table_url,
            )
        )

    return holdings


# =========================
# Cheap holdings aggregation/filter/report helpers
# =========================

def aggregate_holdings(holdings: List[Holding]) -> List[AggregatedHolding]:
    grouped: Dict[Tuple[str, str, str, str], List[Holding]] = {}

    for holding in holdings:
        key = (
            holding.manager_cik,
            holding.report_period,
            holding.cusip,
            holding.title_of_class,
        )
        grouped.setdefault(key, []).append(holding)

    aggregated: List[AggregatedHolding] = []

    for _, rows in grouped.items():
        first = rows[0]
        total_value = sum(row.value or 0.0 for row in rows)
        total_shares = sum(row.shares_or_prn_amount or 0.0 for row in rows)
        implied_price = compute_implied_price(total_value, total_shares)

        aggregated.append(
            AggregatedHolding(
                manager_name=first.manager_name,
                manager_cik=first.manager_cik,
                report_period=first.report_period,
                filing_date=first.filing_date,
                issuer=first.issuer,
                title_of_class=first.title_of_class,
                cusip=first.cusip,
                total_value=total_value,
                total_shares=total_shares,
                implied_price=implied_price,
                row_count=len(rows),
            )
        )

    aggregated.sort(
        key=lambda x: (
            x.manager_name.lower(),
            x.report_period,
            x.implied_price if x.implied_price is not None else float("inf"),
            x.issuer.lower(),
        )
    )
    return aggregated


def filter_cheap_holdings(
    holdings: List[AggregatedHolding],
    max_price: float = 25.0,
    min_total_shares: float = 1.0,
) -> List[AggregatedHolding]:
    filtered: List[AggregatedHolding] = []

    for holding in holdings:
        if holding.implied_price is None:
            continue
        if holding.implied_price > max_price:
            continue
        if holding.total_shares < min_total_shares:
            continue
        filtered.append(holding)

    filtered.sort(
        key=lambda x: (
            x.implied_price if x.implied_price is not None else float("inf"),
            -x.total_value,
            x.issuer.lower(),
        )
    )
    return filtered


def print_cheap_holdings_report(holdings: List[AggregatedHolding], max_rows: int = 25) -> None:
    if not holdings:
        print("\nNo cheap holdings matched the current filter.\n")
        return

    print("\nCheap holdings report")
    print("-" * 110)
    print(f"{'Manager':25} {'Issuer':28} {'CUSIP':10} {'Implied $':>10} {'Shares':>14} {'Value':>14} {'Rows':>5}")
    print("-" * 110)

    for holding in holdings[:max_rows]:
        implied_price_text = f"{holding.implied_price:.2f}" if holding.implied_price is not None else "N/A"
        print(
            f"{holding.manager_name[:25]:25} "
            f"{holding.issuer[:28]:28} "
            f"{holding.cusip[:10]:10} "
            f"{implied_price_text:>10} "
            f"{holding.total_shares:14.0f} "
            f"{holding.total_value:14.0f} "
            f"{holding.row_count:5d}"
        )

    print("-" * 110)
    print(f"Showing {min(len(holdings), max_rows)} of {len(holdings)} matching aggregated holdings.\n")


# =========================
# Excel export helpers
# =========================

def autosize_worksheet(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


def export_to_excel(
    raw_holdings: List[Holding],
    aggregated_holdings: List[AggregatedHolding],
    output_path: str,
) -> None:
    wb = Workbook()

    controls_ws = wb.active
    controls_ws.title = "Controls"
    controls_ws.append(["Setting", "Value"])
    controls_ws.append(["Max Implied Price", 25])
    controls_ws.append(["Min Total Shares", 1])
    controls_ws.append(["Min Total Value", 0])
    controls_ws.append(["Last Refresh", time.strftime("%Y-%m-%d %H:%M:%S")])

    raw_ws = wb.create_sheet("Raw_Holdings")
    raw_ws.append([
        "Manager", "Manager CIK", "Filing Date", "Report Period", "Issuer", "Class",
        "CUSIP", "Value", "Shares", "Shares Type", "Investment Discretion",
        "Filing Detail URL", "Info Table URL",
    ])
    for holding in raw_holdings:
        raw_ws.append([
            holding.manager_name, holding.manager_cik, holding.filing_date, holding.report_period,
            holding.issuer, holding.title_of_class, holding.cusip, holding.value,
            holding.shares_or_prn_amount, holding.shares_or_prn_type, holding.investment_discretion,
            holding.filing_detail_url, holding.info_table_url,
        ])

    agg_ws = wb.create_sheet("Aggregated_Holdings")
    agg_ws.append([
        "Manager", "Manager CIK", "Report Period", "Filing Date", "Issuer", "Class", "CUSIP",
        "Total Value", "Total Shares", "Implied Price", "Row Count", "Current Price",
        "Price Difference", "Matches Filter?",
    ])
    agg_ws.auto_filter.ref = "A1:N1"
    for idx, holding in enumerate(aggregated_holdings, start=2):
        agg_ws.append([
            holding.manager_name,
            holding.manager_cik,
            holding.report_period,
            holding.filing_date,
            holding.issuer,
            holding.title_of_class,
            holding.cusip,
            holding.total_value,
            holding.total_shares,
            holding.implied_price,
            holding.row_count,
            "",
            f'=IFERROR(L{idx}-J{idx},"")',
            f'=IF(AND(J{idx}<=Controls!$B$2,I{idx}>=Controls!$B$3,H{idx}>=Controls!$B$4),"YES","NO")',
        ])

    watch_ws = wb.create_sheet("Watchlist")
    watch_ws["A1"] = "Use the Aggregated_Holdings sheet for live filtering."
    watch_ws["A2"] = "The 'Matches Filter?' column there updates automatically from the Controls sheet."
    watch_ws["A3"] = "Then use Excel's filter on the header row and keep only YES values."

    for ws in [controls_ws, raw_ws, agg_ws]:
        autosize_worksheet(ws)
    watch_ws.column_dimensions["A"].width = 80
    wb.save(output_path)


# =========================
# Main workflow
# =========================

def fetch_holdings_for_manager(session: requests.Session, manager_name: str, cik: str, max_filings: int = 2) -> List[Holding]:
    all_holdings: List[Holding] = []

    logging.info("Checking manager: %s (%s)", manager_name, cik)
    detail_urls = get_recent_filing_detail_urls(session, cik, max_filings=max_filings)

    if not detail_urls:
        logging.warning("No filing detail URLs found for %s", manager_name)
        return all_holdings

    for detail_url in detail_urls:
        try:
            filing_info = parse_filing_detail_page(session, detail_url)
            info_table_url = filing_info["info_table_url"]

            if not info_table_url:
                logging.warning("No information table XML found: %s", detail_url)
                continue

            xml_response = safe_get(session, info_table_url)
            print(f"Fetched URL: {xml_response.url}")
            print(f"Content-Type: {xml_response.headers.get('Content-Type', '')}\n")
            holdings = parse_13f_info_table_xml(
                xml_text=xml_response.text,
                manager_name=manager_name,
                manager_cik=cik,
                filing_date=filing_info["filing_date"],
                report_period=filing_info["report_period"],
                detail_url=detail_url,
                info_table_url=info_table_url,
            )
            logging.info("Parsed %s holdings from %s", len(holdings), info_table_url)
            all_holdings.extend(holdings)

        except Exception as exc:
            logging.exception("Failed processing filing %s: %s", detail_url, exc)

    return all_holdings


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )

    session = make_session()
    excel_output_path = str(Path.home() / "Desktop" / "stock_tracker_output.xlsx")
    all_holdings: List[Holding] = []

    for manager in MANAGERS:
        manager_holdings = fetch_holdings_for_manager(
            session=session,
            manager_name=manager["name"],
            cik=manager["cik"],
            max_filings=2,
        )
        all_holdings.extend(manager_holdings)

    print(f"\nTotal raw holdings parsed: {len(all_holdings)}")

    aggregated_holdings = aggregate_holdings(all_holdings)
    print(f"Total aggregated holdings: {len(aggregated_holdings)}")

    cheap_holdings = filter_cheap_holdings(
        aggregated_holdings,
        max_price=25.0,
        min_total_shares=1.0,
    )
    print(f"Total cheap aggregated holdings (<= $25 implied price): {len(cheap_holdings)}")

    print_cheap_holdings_report(cheap_holdings, max_rows=25)

    export_to_excel(
        raw_holdings=all_holdings,
        aggregated_holdings=aggregated_holdings,
        output_path=excel_output_path,
    )
    print(f"Excel workbook written to: {excel_output_path}\n")

    print("Sample aggregated rows:")
    for holding in aggregated_holdings[:10]:
        print(asdict(holding))


if __name__ == "__main__":
    main()
